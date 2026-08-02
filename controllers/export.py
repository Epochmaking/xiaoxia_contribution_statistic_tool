import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from PySide6.QtCore import Qt

from sqlalchemy import select

from database.db import get_session
from models.ui_models import ArticleListViewModel
from models.article_models import Creator, Article, ArticleCreatorLink

def export_to_file(folder_path: Path, to_calc_fee: bool = False):
    """
    导出文章列表到Excel文件
    """
    fee_list_path: Path | None = None
    article_list_path = folder_path / f"作品清单_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    article_model = ArticleListViewModel(to_calc_fee=to_calc_fee)
    column_order = article_model.column_order

    row_cnt = article_model.row_count
    col_cnt = article_model.column_count

    wb = Workbook()
    ws = wb.active

    if ws is None:
        raise ValueError("ws is None")

    # 通用边框
    thin = Side(style="thin", color="cccccc")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # 表头样式：加粗、居中、浅蓝底色
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", start_color="DCE6F1")
    # 内容单元格居中并自动换行
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写入表头
    for c in range(col_cnt):
        cell = ws.cell(row=1, column=c+1, value=article_model.headerData(c, Qt.Orientation.Horizontal))
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        cell.fill = header_fill

    # 写入表格数据
    for r in range(row_cnt):
        for c in range(col_cnt):
            key = column_order[c]

            if key == "content_url": # 内容链接需要写入完整链接
                val = article_model.index(r, c).data(role = Qt.ItemDataRole.UserRole)
            else:
                val = article_model.index(r, c).data()

            cell = ws.cell(row=r+2, column=c+1, value=val)
            cell.alignment = cell_align
            cell.border = border

    # 自适应列宽简易处理
    for col in range(1, col_cnt+1):
        ws.column_dimensions[chr(64+col)].width = 20

    wb.save(article_list_path)

    if to_calc_fee:
        fee_list_path = folder_path / f"稿费清单_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        fee_wb = Workbook()
        ws_xiaoxia = fee_wb.active
        if ws_xiaoxia is None:
            raise ValueError("ws_xiaoxia is None")
        ws_xiaoxia.title = "小夏成员"
        ws_non_xiaoxia: Worksheet | None = fee_wb.create_sheet(title="非小夏成员")
        if ws_non_xiaoxia is None:
            raise ValueError("ws_non_xiaoxia is None")

        fee_headers = ["序号", "姓名", "参与作品", "稿费（元）"]

        for sheet in [ws_xiaoxia, ws_non_xiaoxia]:
            for c, header in enumerate(fee_headers):
                cell = sheet.cell(row=1, column=c+1, value=header)
                cell.font = header_font
                cell.alignment = header_align
                cell.border = border
                cell.fill = header_fill

        with get_session() as session:
            articles_in_order = session.query(Article).order_by(Article.id.asc()).all()
            article_id_to_seq_title = {article.id: (i + 1, article.title) for i, article in enumerate(articles_in_order)}

            creators = session.query(Creator).order_by(Creator.fee.desc()).all()

            creator_article_ids = {}
            for creator in creators:
                links = session.execute(
                    select(ArticleCreatorLink.article_id).where(
                        ArticleCreatorLink.creator_id == creator.id
                    )
                ).all()
                creator_article_ids[creator.id] = [link[0] for link in links]

        xiaoxia_row = 2
        non_xiaoxia_row = 2
        xiaoxia_index = 1
        non_xiaoxia_index = 1

        for creator in creators:
            if creator.is_xiaoxia:
                sheet = ws_xiaoxia
                row = xiaoxia_row
                idx = xiaoxia_index
                xiaoxia_row += 1
                xiaoxia_index += 1
            else:
                sheet = ws_non_xiaoxia
                row = non_xiaoxia_row
                idx = non_xiaoxia_index
                non_xiaoxia_row += 1
                non_xiaoxia_index += 1

            ordered = sorted(
                (article_id_to_seq_title[aid] for aid in creator_article_ids[creator.id] if aid in article_id_to_seq_title),
                key=lambda x: x[0]
            )
            involved_articles = "\n".join(f"{i + 1}.《{title}》" for i, (_, title) in enumerate(ordered))

            values = [idx, creator.name, involved_articles, creator.fee]
            for c, val in enumerate(values):
                cell = sheet.cell(row=row, column=c+1, value=val)
                cell.alignment = cell_align
                cell.border = border

        sheets: list[Worksheet] | None = [ws_xiaoxia, ws_non_xiaoxia]
        for sheet in sheets:
            for col in range(1, len(fee_headers) + 1):
                sheet.column_dimensions[chr(64 + col)].width = 20

        fee_wb.save(fee_list_path)